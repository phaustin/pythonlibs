from openpyxl import load_workbook
import pandas as pd

def cleanit(item):
        cleanitem = item.encode('ascii', 'ignore').decode('ascii')
        return cleanitem
    

def make_simple(the_file,numcols=None):
    wb=load_workbook(the_file,data_only=True,use_iterators=True)
    combine,=wb.get_sheet_names()
    sheet=wb[combine]
    row_iter=sheet.iter_rows()
    headers = [c.value for c in next(row_iter) if c.value]
    if numcols:
        endcol=numcols
    else:
        endcol=len(headers)
    lines=[]
    for count,row in enumerate(row_iter):
        #print('reading row: ',count)
        values=[item.value for item in row]
        values=values[:endcol]
        #print('values are: ',values)
        #rint('row dump: ',dict(zip(header_list,values)))
        lines.append(values[:endcol])
    clean_head=[cleanit(item) for item in headers[:endcol]]
    df_out=pd.DataFrame.from_records(lines,columns=clean_head)
    # test=df_out.loc[:20]
    # for item in test:
    #     youritem = item.encode('ascii', 'ignore').decode('ascii')
    # yourstring = test.encode('ascii', 'ignore').decode('ascii')
    # print(df_out.head())
    return df_out
