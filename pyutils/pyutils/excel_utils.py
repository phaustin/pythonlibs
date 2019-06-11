from openpyxl import load_workbook
import pdb
import pandas as pd
from collections import OrderedDict as od
import numpy as np,pdb
from pathlib import Path

def cleanit(item):
        cleanitem = item.encode('ascii', 'ignore').decode('ascii')
        return cleanitem
    

def make_simple(the_file,numcols=None):
    the_file=Path(the_file)    
    wb=load_workbook(str(the_file),data_only=True)
    combine,=wb.get_sheet_names()
    sheet=wb[combine]
    row_iter=sheet.iter_rows()
    headers = [c.value for c in next(row_iter) if c.value]
    if numcols:
        endcol=numcols
    else:
        endcol=len(headers)
    clean_head=[cleanit(item) for item in headers[:endcol]]
    lines=[]
    for count,row in enumerate(row_iter):
        #pdb.set_trace()
        values=[item.value for item in row]
        #print("values: ",values)
        values=values[:endcol]
        clean_values = []
        for name, value in zip(clean_head,values):
            if (name == 'Username' or name == 'Student Number' or
                name == "Surname" or name == "Given Name"):
                clean_values.append(str(value))
            else:
                try:
                    clean_values.append(float(value))
                except (ValueError,TypeError):
                    clean_values.append(value)
        lines.append(clean_values)
    df_out=pd.DataFrame.from_records(lines,columns=clean_head)
    #pdb.set_trace()
    # test=df_out.loc[:20]
    # for item in test:
    #     youritem = item.encode('ascii', 'ignore').decode('ascii')
    # yourstring = test.encode('ascii', 'ignore').decode('ascii')
    # print(df_out.head())
    return df_out


def make_simple_headers(the_file,numcols=None):
    wb=load_workbook(the_file,data_only=True,read_only=True)
    combine,=wb.get_sheet_names()
    sheet=wb[combine]
    row_iter=sheet.iter_rows()
    headers = [c.value for c in next(row_iter) if c.value]
    if numcols:
        endcol=numcols
    else:
        endcol=len(headers)
    clean_head=[cleanit(item) for item in headers[:endcol]]
    head_dict = od()
    [head_dict.__setitem__(k,v) for (k,v) in enumerate(clean_head)]
    lines=[]
    for count,row in enumerate(row_iter):
        #print('reading row: ',count)
        values=[item.value for item in row]
        values=values[:endcol]
        #print('values are: ',values)
        #print('row dump: ',dict(zip(clean_head,values)))
        lines.append(values[:endcol])
    df_out=pd.DataFrame.from_records(lines,columns=head_dict.keys())
    # test=df_out.loc[:20]
    # for item in test:
    #     youritem = item.encode('ascii', 'ignore').decode('ascii')
    # yourstring = test.encode('ascii', 'ignore').decode('ascii')
    # print(df_out.head())
    return head_dict,df_out
